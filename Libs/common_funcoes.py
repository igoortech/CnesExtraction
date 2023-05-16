import time
import os
import shutil
import pandas as pd
from openpyxl import load_workbook


def aguardar_renomeia_move(arquivo,novo_arquivo):
    # Aguarda a existência do arquivo
    while not os.path.exists(arquivo):
        print(f"O arquivo '{arquivo}' não existe ainda. Aguardando...")
        time.sleep(5)  # Pausa a execução do programa por 5 segundos
    time.sleep(2)
    print("baixou")
    shutil.copy2(arquivo, novo_arquivo) 
    os.remove(arquivo)



def gerar_excel_output(dados,caminho_arquivo):        
        # Verifica se o arquivo Excel existe
    if os.path.isfile(caminho_arquivo):
        # Carrega o arquivo Excel existente
        df = pd.read_excel(caminho_arquivo)
    else:
        # Cria um DataFrame vazio
        df = pd.DataFrame()

    # Cria um DataFrame com os novos dados
    df_novos = pd.DataFrame(dados, columns=['uf','municipio','Cnes','Nome_Fantasia','Natureza','Gestao','Atende_Sus','url_ficha','STATUS'])
    
    # Concatena o DataFrame existente com os novos dados
    df_concatenado = pd.concat([df, df_novos], ignore_index=True)

    # Salva o DataFrame concatenado no arquivo Excel
    df_concatenado.to_excel(caminho_arquivo, index=False, sheet_name="Dados Consolidados", )



def set_cell_sheet_index(arquivo: str, sheet, celula: str, valor: str):
    wb = load_workbook(arquivo)
    ws = wb.worksheets[sheet]
    ws[celula] = valor
    wb.save(arquivo)



